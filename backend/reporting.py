import csv
import json
import os
from datetime import datetime, timedelta
from io import BytesIO, StringIO

try:
    from .database import Alert, ParkingHistory, ParkingSlot, Transaction
except ImportError:
    from database import Alert, ParkingHistory, ParkingSlot, Transaction


def _format_datetime(value):
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else "N/A"


def _format_duration(minutes):
    if minutes is None:
        return "N/A"
    minutes = int(minutes)
    hours, mins = divmod(minutes, 60)
    if hours and mins:
        return f"{hours}h {mins}m"
    if hours:
        return f"{hours}h"
    return f"{mins}m"


def _format_currency(amount):
    return f"${float(amount or 0):.2f}"


def _humanize_alert_type(alert_type):
    return (alert_type or "unknown").replace("_", " ").title()


def _format_alert_detail(detail):
    if not detail:
        return "N/A"
    if isinstance(detail, dict):
        return ", ".join(f"{k}: {v}" for k, v in detail.items()) or "N/A"
    try:
        parsed = json.loads(detail)
        if isinstance(parsed, dict):
            return ", ".join(f"{k}: {v}" for k, v in parsed.items()) or "N/A"
        return str(parsed)
    except Exception:
        return str(detail)


def collect_report_data(db, recent_limit=10, activity_limit=15, alert_limit=10):
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = now - timedelta(days=now.weekday())
    week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    slots = db.query(ParkingSlot).all()
    total_slots = len(slots)
    available = sum(1 for slot in slots if slot.status == "available")
    occupied = total_slots - available
    occupancy_pct = round((occupied / total_slots) * 100, 1) if total_slots > 0 else 0

    def _sum_amount(query_rows):
        return round(sum(row[0] or 0 for row in query_rows), 2)

    today_rev = _sum_amount(
        db.query(Transaction.amount)
        .filter(Transaction.entry_time >= today_start, Transaction.status == "completed")
        .all()
    )
    week_rev = _sum_amount(
        db.query(Transaction.amount)
        .filter(Transaction.entry_time >= week_start, Transaction.status == "completed")
        .all()
    )
    month_rev = _sum_amount(
        db.query(Transaction.amount)
        .filter(Transaction.entry_time >= month_start, Transaction.status == "completed")
        .all()
    )

    completed_tx = db.query(Transaction).filter(Transaction.status == "completed").all()
    avg_per_vehicle = round(
        sum(t.amount or 0 for t in completed_tx) / len(completed_tx), 2
    ) if completed_tx else 0

    recent_tx = (
        db.query(Transaction)
        .order_by(Transaction.entry_time.desc())
        .limit(recent_limit)
        .all()
    )
    recent_events = (
        db.query(ParkingHistory)
        .order_by(ParkingHistory.timestamp.desc())
        .limit(activity_limit)
        .all()
    )
    active_alerts = (
        db.query(Alert)
        .filter(Alert.resolved == False)
        .order_by(Alert.timestamp.desc())
        .limit(alert_limit)
        .all()
    )

    return {
        "now": now,
        "date_str": now.strftime("%Y-%m-%d"),
        "today_str": now.strftime("%A, %B %d, %Y"),
        "generated_at": _format_datetime(now),
        "overview": {
            "total_slots": total_slots,
            "available": available,
            "occupied": occupied,
            "occupancy_pct": occupancy_pct,
        },
        "revenue": {
            "today": today_rev,
            "week": week_rev,
            "month": month_rev,
            "avg_per_vehicle": avg_per_vehicle,
        },
        "recent_transactions": [
            {
                "time": _format_datetime(item.entry_time),
                "plate": item.plate or "N/A",
                "vehicle_type": item.vehicle_type or "N/A",
                "slot": item.slot_id or "N/A",
                "duration": _format_duration(item.duration_mins),
                "amount": _format_currency(item.amount),
                "status": (item.status or "N/A").title(),
            }
            for item in recent_tx
        ],
        "recent_activity": [
            {
                "time": _format_datetime(item.timestamp),
                "slot": item.slot_id or "N/A",
                "event": "Entry" if item.status == "occupied" else "Exit",
                "duration": _format_duration(item.dwell_minutes),
                "vehicle_type": item.vehicle_type or "N/A",
                "plate": item.plate or "N/A",
            }
            for item in recent_events
        ],
        "active_alerts": [
            {
                "time": _format_datetime(item.timestamp),
                "type": _humanize_alert_type(item.alert_type),
                "slot": item.slot_id or "N/A",
                "vehicle": item.vehicle_id or "N/A",
                "detail": _format_alert_detail(item.detail),
                "resolved": bool(item.resolved),
            }
            for item in active_alerts
        ],
    }


def _write_csv_rows(writer, report):
    writer.writerow(["Smart Parking Report"])
    writer.writerow(["Generated", report["generated_at"]])
    writer.writerow(["Date", report["today_str"]])
    writer.writerow([])

    writer.writerow(["Overview"])
    writer.writerow(["Total Slots", "Available", "Occupied", "Occupancy %"])
    writer.writerow([
        report["overview"]["total_slots"],
        report["overview"]["available"],
        report["overview"]["occupied"],
        f'{report["overview"]["occupancy_pct"]}%',
    ])
    writer.writerow([])

    writer.writerow(["Revenue Summary"])
    writer.writerow(["Today", "This Week", "This Month", "Avg Per Vehicle"])
    writer.writerow([
        _format_currency(report["revenue"]["today"]),
        _format_currency(report["revenue"]["week"]),
        _format_currency(report["revenue"]["month"]),
        _format_currency(report["revenue"]["avg_per_vehicle"]),
    ])
    writer.writerow([])

    writer.writerow(["Recent Transactions"])
    writer.writerow(["Time", "Plate", "Vehicle Type", "Slot", "Duration", "Amount", "Status"])
    if report["recent_transactions"]:
        for row in report["recent_transactions"]:
            writer.writerow([
                row["time"], row["plate"], row["vehicle_type"], row["slot"],
                row["duration"], row["amount"], row["status"],
            ])
    else:
        writer.writerow(["No recent transactions"])
    writer.writerow([])

    writer.writerow(["Recent Activity"])
    writer.writerow(["Time", "Slot", "Event", "Duration", "Vehicle Type", "Plate"])
    if report["recent_activity"]:
        for row in report["recent_activity"]:
            writer.writerow([
                row["time"], row["slot"], row["event"], row["duration"],
                row["vehicle_type"], row["plate"],
            ])
    else:
        writer.writerow(["No recent activity"])
    writer.writerow([])

    writer.writerow([f'Active Alerts ({len(report["active_alerts"])})'])
    writer.writerow(["Time", "Type", "Slot", "Vehicle", "Detail"])
    if report["active_alerts"]:
        for row in report["active_alerts"]:
            writer.writerow([
                row["time"], row["type"], row["slot"], row["vehicle"], row["detail"],
            ])
    else:
        writer.writerow(["No active alerts"])


def build_csv_report(report):
    buffer = StringIO()
    writer = csv.writer(buffer)
    _write_csv_rows(writer, report)
    return buffer.getvalue().encode("utf-8")


def write_csv_report(filepath, report):
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        _write_csv_rows(writer, report)


def build_pdf_report(report):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    def _styled_table(rows, widths, header_bg="#EFF6FF", body_bg="#FFFFFF", align="LEFT"):
        table = Table(rows, colWidths=widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_bg)),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor(body_bg)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("LEADING", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D7DEE7")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), align),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return table

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontSize=22, textColor=colors.HexColor("#0F172A"), spaceAfter=4)
    subtitle_style = ParagraphStyle("ReportSub", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#64748B"), spaceAfter=8)
    section_style = ParagraphStyle("Section", parent=styles["Heading2"], fontSize=12, textColor=colors.HexColor("#2563EB"), spaceBefore=8, spaceAfter=6)

    story = [
        Paragraph("Smart Parking Report", title_style),
        Paragraph(f'{report["today_str"]} | Generated {report["generated_at"]}', subtitle_style),
        HRFlowable(width="100%", thickness=0.7, color=colors.HexColor("#D7DEE7")),
        Spacer(1, 5 * mm),
    ]

    story.append(Paragraph("Overview", section_style))
    story.append(_styled_table(
        [
            ["Total Slots", "Available", "Occupied", "Occupancy"],
            [
                report["overview"]["total_slots"],
                report["overview"]["available"],
                report["overview"]["occupied"],
                f'{report["overview"]["occupancy_pct"]}%',
            ],
        ],
        [42 * mm, 42 * mm, 42 * mm, 42 * mm],
        header_bg="#E0F2FE",
        body_bg="#F8FAFC",
        align="CENTER",
    ))
    story.append(Spacer(1, 4 * mm))

    story.append(Paragraph("Revenue Summary", section_style))
    story.append(_styled_table(
        [
            ["Today", "This Week", "This Month", "Avg / Vehicle"],
            [
                _format_currency(report["revenue"]["today"]),
                _format_currency(report["revenue"]["week"]),
                _format_currency(report["revenue"]["month"]),
                _format_currency(report["revenue"]["avg_per_vehicle"]),
            ],
        ],
        [42 * mm, 42 * mm, 42 * mm, 42 * mm],
        header_bg="#DCFCE7",
        body_bg="#F8FAFC",
        align="CENTER",
    ))
    story.append(Spacer(1, 4 * mm))

    tx_rows = [["Time", "Plate", "Type", "Slot", "Duration", "Amount", "Status"]]
    if report["recent_transactions"]:
        for row in report["recent_transactions"]:
            tx_rows.append([
                row["time"], row["plate"], row["vehicle_type"], row["slot"],
                row["duration"], row["amount"], row["status"],
            ])
    else:
        tx_rows.append(["No recent transactions", "", "", "", "", "", ""])
    story.append(Paragraph("Recent Transactions", section_style))
    story.append(_styled_table(tx_rows, [34 * mm, 24 * mm, 23 * mm, 18 * mm, 22 * mm, 18 * mm, 18 * mm]))
    story.append(Spacer(1, 4 * mm))

    activity_rows = [["Time", "Slot", "Event", "Duration", "Type", "Plate"]]
    if report["recent_activity"]:
        for row in report["recent_activity"]:
            activity_rows.append([
                row["time"], row["slot"], row["event"], row["duration"], row["vehicle_type"], row["plate"],
            ])
    else:
        activity_rows.append(["No recent activity", "", "", "", "", ""])
    story.append(Paragraph("Recent Activity", section_style))
    story.append(_styled_table(activity_rows, [34 * mm, 18 * mm, 18 * mm, 22 * mm, 28 * mm, 28 * mm], header_bg="#F1F5F9"))
    story.append(Spacer(1, 4 * mm))

    alert_rows = [["Time", "Type", "Slot", "Vehicle", "Detail"]]
    if report["active_alerts"]:
        for row in report["active_alerts"]:
            alert_rows.append([row["time"], row["type"], row["slot"], row["vehicle"], row["detail"]])
    else:
        alert_rows.append(["No active alerts", "", "", "", ""])
    story.append(Paragraph(f'Active Alerts ({len(report["active_alerts"])})', section_style))
    story.append(_styled_table(alert_rows, [34 * mm, 24 * mm, 16 * mm, 24 * mm, 66 * mm], header_bg="#FEE2E2"))
    story.append(Spacer(1, 6 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#D7DEE7")))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph("Smart Parking System", ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#94A3B8"), alignment=1)))

    doc.build(story)
    buf.seek(0)
    return buf.read()


def build_excel_report(report):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    border = Border(
        left=Side(style="thin", color="D7DEE7"),
        right=Side(style="thin", color="D7DEE7"),
        top=Side(style="thin", color="D7DEE7"),
        bottom=Side(style="thin", color="D7DEE7"),
    )
    header_fill = PatternFill(fill_type="solid", start_color="2563EB", end_color="2563EB")
    alert_fill = PatternFill(fill_type="solid", start_color="FEE2E2", end_color="FEE2E2")
    summary_fill = PatternFill(fill_type="solid", start_color="F8FAFC", end_color="F8FAFC")
    white_bold = Font(color="FFFFFF", bold=True, size=11)

    def _style_header(row, fill=header_fill):
        for cell in row:
            cell.font = white_bold if fill == header_fill else Font(bold=True, size=11, color="0F172A")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def _write_table(ws, start_row, headers, rows, fill=header_fill):
        for col, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=header)
        _style_header(ws[start_row], fill)
        row = start_row + 1
        if not rows:
            ws.cell(row=row, column=1, value="No data")
            ws.cell(row=row, column=1).border = border
            return row
        for item in rows:
            for col, value in enumerate(item, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top")
            row += 1
        return row - 1

    summary = wb.active
    summary.title = "Summary"
    summary.merge_cells("A1:D1")
    summary["A1"] = "Smart Parking Report"
    summary["A1"].font = Font(bold=True, size=18, color="0F172A")
    summary["A2"] = report["today_str"]
    summary["A3"] = f'Generated {report["generated_at"]}'

    _write_table(
        summary,
        5,
        ["Total Slots", "Available", "Occupied", "Occupancy %"],
        [[
            report["overview"]["total_slots"],
            report["overview"]["available"],
            report["overview"]["occupied"],
            report["overview"]["occupancy_pct"],
        ]],
        fill=summary_fill,
    )
    _write_table(
        summary,
        9,
        ["Today", "This Week", "This Month", "Avg / Vehicle"],
        [[
            _format_currency(report["revenue"]["today"]),
            _format_currency(report["revenue"]["week"]),
            _format_currency(report["revenue"]["month"]),
            _format_currency(report["revenue"]["avg_per_vehicle"]),
        ]],
        fill=summary_fill,
    )

    sheets = [
        ("Transactions", ["Time", "Plate", "Type", "Slot", "Duration", "Amount", "Status"], report["recent_transactions"]),
        ("Activity", ["Time", "Slot", "Event", "Duration", "Type", "Plate"], report["recent_activity"]),
        ("Alerts", ["Time", "Type", "Slot", "Vehicle", "Detail"], report["active_alerts"]),
    ]

    for name, headers, rows in sheets:
        ws = wb.create_sheet(name)
        ws["A1"] = name
        ws["A1"].font = Font(bold=True, size=16, color="0F172A")
        row_values = [[item.get(h.lower().replace(" / ", "_").replace(" ", "_"), item.get(h.lower(), "")) for h in headers] for item in rows]
        _write_table(ws, 3, headers, row_values, fill=alert_fill if name == "Alerts" else header_fill)
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{chr(64 + len(headers))}{max(4, len(row_values) + 3)}"
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            ws.column_dimensions[column_letter].width = min(max_length + 4, 36)

    for column in ["A", "B", "C", "D"]:
        summary.column_dimensions[column].width = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def save_report_files(report, export_dir):
    os.makedirs(export_dir, exist_ok=True)
    csv_name = f'report_{report["date_str"]}.csv'
    pdf_name = f'SmartParking_Report_{report["date_str"]}.pdf'
    xlsx_name = f'SmartParking_Report_{report["date_str"]}.xlsx'

    csv_path = os.path.join(export_dir, csv_name)
    pdf_path = os.path.join(export_dir, pdf_name)
    xlsx_path = os.path.join(export_dir, xlsx_name)

    write_csv_report(csv_path, report)
    with open(pdf_path, "wb") as f:
        f.write(build_pdf_report(report))
    with open(xlsx_path, "wb") as f:
        f.write(build_excel_report(report))

    return {
        "csv": {"filename": csv_name, "path": csv_path, "size": os.path.getsize(csv_path)},
        "pdf": {"filename": pdf_name, "path": pdf_path, "size": os.path.getsize(pdf_path)},
        "xlsx": {"filename": xlsx_name, "path": xlsx_path, "size": os.path.getsize(xlsx_path)},
    }
